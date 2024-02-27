#def doprocess_Fresh(INPATH,OUTPATH):
import pandas as pd
import numpy as np
import datetime as date
import shutil
import os
import requests
import time
from openpyxl import load_workbook
import os
import openpyxl
import random  
import pathlib
import openpyxl, time, os, os.path, datetime
from datetime import datetime, timedelta 
from openpyxl.styles import Border, Side, Font, GradientFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import os
import time        
from datetime import date
import datetime
import os.path
from os import path
from pathlib import Path
from datetime import datetime, timedelta 
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os
import glob, shutil
from datetime import datetime, timedelta
import datetime
from datetime import datetime  
from datetime import timedelta 
import openpyxl
from datetime import datetime  
from openpyxl.chart import LineChart,Reference
from openpyxl.chart.axis import DateAxis
from openpyxl import Workbook, chart
import os
from openpyxl.chart import LineChart, Reference, Series
from datetime import date, datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

INPATH='C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data'
INPATH2='C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data'
INPATH1='C://EVENTS_AUTO'
OUTPATH='C://EVENTS_AUTO//Output_ISF'



##################################################################Remove old Raw CSV files inside ITK_Raw_data folder##################################
'''file = Path("C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Cell_New.csv")
if file.exists ():
        print ("2G_Raw_Cell_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Cell_New.csv")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Total_New.csv")
if file.exists ():
        print ("2G_Raw_Total_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Total_New.csv")
else:
        print ("File not exist")
                
file = Path("C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Cell_New.csv")
if file.exists ():
        print ("3G_Raw_Cell_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Cell_New.csv")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Total_New.csv")
if file.exists ():
        print ("3G_Raw_Total_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Total_New.csv")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Cell_New.csv")
if file.exists ():
        print ("4G_Raw_Cell_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Cell_New.csv")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Total_New.csv")
if file.exists ():
        print ("4G_Raw_Total_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Total_New.csv")
else:
        print ("File not exist")


file = Path("C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Cell_New.csv")
if file.exists ():
        print ("5G_Raw_Cell_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Cell_New.csv")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Total_New.csv")
if file.exists ():
        print ("5G_Raw_Total_New File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Total_New.csv")
else:
        print ("File not exist")'''

#####################################################################Move Fresh Raw data inside ITK_Raw_data folder#########################################
'''file = Path("C://EVENTS_AUTO//one.csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Cell_New.csv')

file = Path("C://EVENTS_AUTO//one (1).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Total_New.csv')

file = Path("C://EVENTS_AUTO//one (2).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Cell_New.csv')

file = Path("C://EVENTS_AUTO//one (3).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Total_New.csv')

file = Path("C://EVENTS_AUTO//one (4).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Cell_New.csv')

file = Path("C://EVENTS_AUTO//one (5).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Total_New.csv')


file = Path("C://EVENTS_AUTO//one (6).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Cell_New.csv')

file = Path("C://EVENTS_AUTO//one (7).csv")
if file.is_file():
    print ("one File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Total_New.csv')

################################################################################Delete old file as it is completed#####################################
file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Cell_Old.xlsx")
if file.exists ():
        print ("2G_Raw_Cell_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Cell_Old.xlsx")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Total_Old.xlsx")
if file.exists ():
        print ("2G_Raw_Total_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Total_Old.xlsx")
else:
        print ("File not exist")
                
file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Cell_Old.xlsx")
if file.exists ():
        print ("3G_Raw_Cell_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Cell_Old.xlsx")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Total_Old.xlsx")
if file.exists ():
        print ("3G_Raw_Total_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Total_Old.xlsx")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Cell_Old.xlsx")
if file.exists ():
        print ("4G_Raw_Cell_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Cell_Old.xlsx")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Total_Old.xlsx")
if file.exists ():
        print ("4G_Raw_Total_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Total_Old.xlsx")
else:
        print ("File not exist")


file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Cell_Old.xlsx")
if file.exists ():
        print ("5G_Raw_Cell_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Cell_Old.xlsx")
else:
        print ("File not exist")

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Total_Old.xlsx")
if file.exists ():
        print ("5G_Raw_Total_Old File exist")
        os.remove("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Total_Old.xlsx")
else:
        print ("File not exist")'''


#############################################################Edit the Fresh Raw data compatible to Final output with ehader change##############################
############################################################Edit the ITK Raw files with Optima names###########################################
############################################################Edit the 4ITK Raw files with Optima names###########################################
############################################################Edit the 4ITK Raw files with Optima names###########################################
############################################################Edit the 4ITK Raw files with Optima names###########################################
df4g_kpi=pd.read_excel(INPATH1 +"//Events_Input_Template.xlsx",sheet_name='4G_ITKVsoptima')
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Cell_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df4g_kpi)):
    KPIname_4g_fin=df4g_kpi.loc[i, "4G Final Name"]
    KPIname_4g_itk=df4g_kpi.loc[i, "4G ITK Name"]
    db.rename(columns={KPIname_4g_itk:KPIname_4g_fin},inplace=True)
    db[KPIname_4g_fin]=pd.to_numeric(db[KPIname_4g_fin], errors ='coerce')
db.to_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)
###############Rearrange the file as per final output sheet##########################

df = db[['time','object','4G_ESTABLISHMENT_SUCCESS_RATE','4G_CALL_DROP_RATE','4G_DL_DATA_VOLUME_GB','4G_UL_DATA_VOLUME_GB','AVG_DL_USER_THP_Mbps','AVG_UL_USER_THP_Mbps','AVG_DATA_USER_NUM_DL','AVG_DL_CELL_THP_Mbps','AVG_UL_CELL_THP_Mbps','AVG_DATA_USER_NUM_UL','CONNECTED_USERS','VOLTE_TRAFFIC_ERLANG','CSSR_VOLTE','VOLTE_CDR']]
df1 = db[['time','object','PUCCH_RSSI','PUSCH_RSSI']]
df2 = db[['time','object','4G_VoLTE_SRVCC_HOSR_overall']]


with pd.ExcelWriter(INPATH2 +"//4G_Raw_Cell_New.xlsx") as writer:
      
    df.to_excel(writer,sheet_name='EVENT_4G_KPI',float_format="%.2f",index=False)
    df1.to_excel(writer,sheet_name='EVENT_4G_RSSI',float_format="%.2f",index=False)
    df2.to_excel(writer,sheet_name='EVENT_4G_SRVCC_KPI',float_format="%.2f",index=False)
writer.save() 
################################################Edit for 4G Total Raw data sheet######################################################################
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//4G_Raw_Total_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df4g_kpi)):
    KPIname_4g_fin=df4g_kpi.loc[i, "4G Final Name"]
    KPIname_4g_itk=df4g_kpi.loc[i, "4G ITK Name"]
    db.rename(columns={KPIname_4g_itk:KPIname_4g_fin},inplace=True)
    db[KPIname_4g_fin]=pd.to_numeric(db[KPIname_4g_fin], errors ='coerce')
df = db[['time','4G_ESTABLISHMENT_SUCCESS_RATE','4G_CALL_DROP_RATE','4G_DL_DATA_VOLUME_GB','4G_UL_DATA_VOLUME_GB','AVG_DL_USER_THP_Mbps','AVG_UL_USER_THP_Mbps','AVG_DATA_USER_NUM_DL','AVG_DL_CELL_THP_Mbps','AVG_UL_CELL_THP_Mbps','AVG_DATA_USER_NUM_UL','CONNECTED_USERS','VOLTE_TRAFFIC_ERLANG','CSSR_VOLTE','VOLTE_CDR']]
df1 = db[['time','PUCCH_RSSI','PUSCH_RSSI']]
df2 = db[['time','4G_VoLTE_SRVCC_HOSR_overall']]

with pd.ExcelWriter(INPATH2 +"//4G_Raw_Total_New.xlsx") as writer:
      
    df.to_excel(writer,sheet_name='EVENT_4G_KPI',float_format="%.2f",index=False)
    df1.to_excel(writer,sheet_name='EVENT_4G_RSSI',float_format="%.2f",index=False)
    df2.to_excel(writer,sheet_name='EVENT_4G_SRVCC_KPI',float_format="%.2f",index=False)
writer.save() 




############################################################Edit the 3ITK Raw files with Optima names###########################################
############################################################Edit the 3ITK Raw files with Optima names###########################################
############################################################Edit the 3ITK Raw files with Optima names###########################################
df3g_kpi=pd.read_excel(INPATH1 +"//Events_Input_Template.xlsx",sheet_name='3G_ITKVsoptima')
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Cell_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df3g_kpi)):
    KPIname_3g_fin=df3g_kpi.loc[i, "3G Final Name"]
    KPIname_3g_itk=df3g_kpi.loc[i, "3G ITK Name"]
    db.rename(columns={KPIname_3g_itk:KPIname_3g_fin},inplace=True)
    db[KPIname_3g_fin]=pd.to_numeric(db[KPIname_3g_fin], errors ='coerce')

df = db[['time','object','3G_CS_CSSR','3G_TRAFIK_VOLUME_ERLANG','3G_PS_CSSR','3G_DATA_UL_MB','3G_DATA_DL_MB','3G_CDR','MEAN_RTWP','3G_USER_HSDPA_THROUGHPUT_Mbps','3G_USER_HSUPA_THROUGHPUT_Mbps','FACH_USER_Number','DCH_USER_Number']]

with pd.ExcelWriter(INPATH2 +"//3G_Raw_Cell_New.xlsx") as writer:
      
    df.to_excel(writer,sheet_name='CELL_3G_KPI',float_format="%.2f",index=False)
writer.save() 
################################################Edit for 3G Total Raw data sheet######################################################################
df3g_kpi=pd.read_excel(INPATH1 +"//Events_Input_Template.xlsx",sheet_name='3G_Total')
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//3G_Raw_Total_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df3g_kpi)):
    KPIname_3g_fin=df3g_kpi.loc[i, "3G Final Name_Total"]
    KPIname_3g_itk=df3g_kpi.loc[i, "3G ITK Name_Total"]
    db.rename(columns={KPIname_3g_itk:KPIname_3g_fin},inplace=True)
    db[KPIname_3g_fin]=pd.to_numeric(db[KPIname_3g_fin], errors ='coerce')
###############Rearrange the file as per final output sheet##########################

df = db[['time','3G_CS_CSSR','3G_TRAFIK_VOLUME_ERLANG','3G_PS_CSSR','3G_DATA_UL_MB','3G_DATA_DL_MB','3G_CDR','MEAN_RTWP','3G_USER_HSDPA_THROUGHPUT_Mbps','3G_USER_HSUPA_THROUGHPUT_Mbps']]


with pd.ExcelWriter(INPATH2 +"//3G_Raw_Total_New.xlsx") as writer:
    df.to_excel(writer,sheet_name='CELL_3G_KPI',float_format="%.2f",index=False)
writer.save() 





############################################################Edit the 2ITK Raw files with Optima names###########################################
############################################################Edit the 2ITK Raw files with Optima names###########################################
############################################################Edit the 2ITK Raw files with Optima names###########################################
df2g_kpi=pd.read_excel(INPATH1 +"//Events_Input_Template.xlsx",sheet_name='2G_ITKVsoptima')
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Cell_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df2g_kpi)):
    KPIname_2g_fin=df2g_kpi.loc[i, "2G Final Name"]
    KPIname_2g_itk=df2g_kpi.loc[i, "2G ITK Name"]
    db.rename(columns={KPIname_2g_itk:KPIname_2g_fin},inplace=True)
    db[KPIname_2g_fin]=pd.to_numeric(db[KPIname_2g_fin], errors ='coerce')    
###############Rearrange the file as per final output sheet##########################

df = db[['time','object','GSM_TRAFIK_VOLUME_ERLANG','GSM_CDR','BSS_Call_Estab_SR','IMMEDIATE_ASSIGNMENT_SR','ASSIGNMENT_SR','SDCCH_CDR','GSM_SDCCH_SETUP_SR','GSM_TCH_SETUP_SR','CS_PAGING_DISCARD_RATE']]

with pd.ExcelWriter(INPATH2 +"//2G_Raw_Cell_New.xlsx") as writer:
      
    df.to_excel(writer,sheet_name='CELL_2G_KPI',float_format="%.2f",index=False)
writer.save() 
################################################Edit for 2G Total Raw data sheet######################################################################
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//2G_Raw_Total_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df2g_kpi)):
    KPIname_2g_fin=df2g_kpi.loc[i, "2G Final Name"]
    KPIname_2g_itk=df2g_kpi.loc[i, "2G ITK Name"]
    db.rename(columns={KPIname_2g_itk:KPIname_2g_fin},inplace=True)
    db[KPIname_2g_fin]=pd.to_numeric(db[KPIname_2g_fin], errors ='coerce')
###############Rearrange the file as per final output sheet##########################

df = db[['time','GSM_TRAFIK_VOLUME_ERLANG','GSM_CDR','BSS_Call_Estab_SR','IMMEDIATE_ASSIGNMENT_SR','ASSIGNMENT_SR','SDCCH_CDR','GSM_SDCCH_SETUP_SR','GSM_TCH_SETUP_SR','CS_PAGING_DISCARD_RATE']]

with pd.ExcelWriter(INPATH2 +"//2G_Raw_Total_New.xlsx") as writer:
    df.to_excel(writer,sheet_name='CELL_2G_KPI',float_format="%.2f",index=False)
writer.save()




############################################################Edit the 5ITK Raw files with Optima names###########################################
############################################################Edit the 5ITK Raw files with Optima names###########################################
############################################################Edit the 5ITK Raw files with Optima names###########################################
df5g_kpi=pd.read_excel(INPATH1 +"//Events_Input_Template.xlsx",sheet_name='5G_ITKVsoptima')
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Cell_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df5g_kpi)):
    KPIname_5g_fin=df5g_kpi.loc[i, "5G Final Name"]
    print(KPIname_5g_fin)
    KPIname_5g_itk=df5g_kpi.loc[i, "5G ITK Name"]
    print(KPIname_5g_itk)    
    db.rename(columns={KPIname_5g_itk:KPIname_5g_fin},inplace=True)
    db[KPIname_5g_fin]=pd.to_numeric(db[KPIname_5g_fin], errors ='coerce')

df = db[['time','object','5G_Availability','5G_EN_DC_Est_SR','5G_EN_DC_CDR','5G_MAC_Traffic_Volume_GB','5G_Average_Uplink_Interference_dBm','5G_Average_User_DL_Thp_Mbps','5G_Average_User_UL_Thp_Mbps','5G_DL_MAC_Traffic_Volume_GB','5G_UL_MAC_Traffic_Volume_GB']]
print('done')
with pd.ExcelWriter(INPATH2 +"//5G_Raw_Cell_New.xlsx") as writer:      
    df.to_excel(writer,sheet_name='CELL_5G_KPI',float_format="%.2f",index=False)
writer.save() 
################################################Edit for 5G Total Raw data sheet######################################################################
df5g_kpi=pd.read_excel(INPATH1 +"//Events_Input_Template.xlsx",sheet_name='5G_ITKVsoptima')
db=pd.read_csv("C://EVENTS_AUTO//ITK_Raw_data//5G_Raw_Total_New.csv")
pd.options.display.float_format = '{:.2f}'.format

for i in range(len(df5g_kpi)):
    KPIname_5g_fin=df5g_kpi.loc[i, "5G Final Name"]
    KPIname_5g_itk=df5g_kpi.loc[i, "5G ITK Name"]
    db.rename(columns={KPIname_5g_itk:KPIname_5g_fin},inplace=True)
    db[KPIname_5g_fin]=pd.to_numeric(db[KPIname_5g_fin], errors ='coerce')
###############Rearrange the file as per final output sheet##########################

df = db[['time','5G_Availability','5G_EN_DC_Est_SR','5G_EN_DC_CDR','5G_MAC_Traffic_Volume_GB','5G_Average_Uplink_Interference_dBm','5G_Average_User_DL_Thp_Mbps','5G_Average_User_UL_Thp_Mbps','5G_DL_MAC_Traffic_Volume_GB','5G_UL_MAC_Traffic_Volume_GB']]


with pd.ExcelWriter(INPATH2 +"//5G_Raw_Total_New.xlsx") as writer:
    df.to_excel(writer,sheet_name='CELL_5G_KPI',float_format="%.2f",index=False)
writer.save() 

print('ITK Files edited as per final output done')



#######################################################################################################################################################
#######################################################################################################################################################
############################################################Copy paste the Raw data to Output file for all Technos and save it#########################
#################################################################Code to copy paste in Final Output sheet-CELL_2G_KPI########################################
#######################################################################################################################################################
#######################################################################################################################################################
#df=pd.read_excel(INPATH2+"//2G_Raw_Cell_Old.xlsx")
df1=pd.read_excel(INPATH2+"//2G_Raw_Cell_New.xlsx")
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//2G_Raw_Cell_New_consider.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 2g][CELL_2G_KPI]############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['CELL_2G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//2G_Raw_Cell_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//2G_Raw_Cell_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["CELL_2G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,12,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,12,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''


template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='CELL_2G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['CELL_2G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(13)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
        
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  

#################################################################Code to copy paste in Final Output sheet-EVENT_2G_KPI########################################

#df=pd.read_excel(INPATH2+"//2G_Raw_Total_Old.xlsx")
df1=pd.read_excel(INPATH2+"//2G_Raw_Total_New.xlsx")
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//2G_Raw_Total_New_consider.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 2g][CELL_2G_KPI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['EVENT_2G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//2G_Raw_Total_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//2G_Raw_Total_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["EVENT_2G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,11,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,11,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='EVENT_2G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['EVENT_2G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(11)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  




##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################
####################################################################2G Pasting in the Output sheet is done################################################
##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################
#######################################################################################################################################################
#######################################################################################################################################################
############################################################Copy paste the Raw data to Output file for all Technos and save it#########################
#################################################################Code to copy paste in Final Output sheet-CELL_3G_KPI########################################
#######################################################################################################################################################
#######################################################################################################################################################


#################################################################Code to copy paste in Final Output sheet-CELL_3G_KPI########################################

#df=pd.read_excel(INPATH2+"//3G_Raw_Cell_Old.xlsx")
df1=pd.read_excel(INPATH2+"//3G_Raw_Cell_New.xlsx")
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//3G_Raw_Cell_New_consider.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 2g][CELL_2G_KPI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['CELL_3G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//3G_Raw_Cell_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//3G_Raw_Cell_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["CELL_3G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,16,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,16,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='CELL_3G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['CELL_3G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(14)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
        
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")    


#################################################################Code to copy paste in Final Output sheet-EVENT_3G_KPI########################################

#df=pd.read_excel(INPATH2+"//3G_Raw_Total_Old.xlsx")
df1=pd.read_excel(INPATH2+"//3G_Raw_Total_New.xlsx")
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//3G_Raw_Total_New_consider.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 3g[CELL_3G_KPI]############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['EVENT_3G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//3G_Raw_Total_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//3G_Raw_Total_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["EVENT_3G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,11,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,11,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='EVENT_3G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['EVENT_3G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(11)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  




##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################
####################################################################3G Pasting in the Output sheet is done################################################
##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################
#######################################################################################################################################################
#######################################################################################################################################################
############################################################Copy paste the Raw data to Output file for all Technos and save it#########################
#################################################################Code to copy paste in Final Output sheet-CELL_4G_KPI########################################
#######################################################################################################################################################
#######################################################################################################################################################
#################################################################Code to copy paste in Final Output sheet-CELL_4G_KPI########################################

#df=pd.read_excel(INPATH2+"//4G_Raw_Cell_Old.xlsx",sheet_name='EVENT_4G_KPI')
df1=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_KPI')
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 2g][CELL_2G_KPI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['CELL_4G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_KPI')
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//4G_Raw_Cell_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
#sheet = wb["EVENT_4G_KPI"] 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["CELL_4G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,18,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,18,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='CELL_4G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['CELL_4G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(17)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
        
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")    



#################################################################Code to copy paste in Final Output sheet-EVENT_4G_KPI########################################

#df=pd.read_excel(INPATH2+"//4G_Raw_Total_Old.xlsx",sheet_name='EVENT_4G_KPI')
df1=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_KPI')
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//4G_Raw_Total_New.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 4g][CELL_4G_KPI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['EVENT_4G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//4G_Raw_Total_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
#sheet = wb["EVENT_4G_KPI"] 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["EVENT_4G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,17,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,17,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='EVENT_4G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['EVENT_4G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(16)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  


file = Path(INPATH2+"//4G_Raw_Total_New.xlsx")


########################################################Copy paste 4G Cell data###################################################################
##################################################################################################################################################
        
#################################################################Code to copy paste in Final Output sheet-EVENT_4G_RSSI########################################

#df=pd.read_excel(INPATH2+"//4G_Raw_Cell_Old.xlsx",sheet_name='EVENT_4G_RSSI')
df1=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_RSSI')
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 4gEVENT_4G_RSSI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['CELL_4G_RSSI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_RSSI')
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//4G_Raw_Cell_New.xlsx") #Add file name
sheet = wb.worksheets[1] #Add Sheet name
#sheet = wb["EVENT_4G_RSSI"] 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["CELL_4G_RSSI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,5,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,5,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='CELL_4G_RSSI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['CELL_4G_RSSI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(5)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
        
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")    



#################################################Copy range and paste in 4g][RSSI_4G_KPI############################################################
#################################################################Code to copy paste in Final Output sheet-Total_4G_RSSI########################################

#df=pd.read_excel(INPATH2+"//4G_Raw_Total_Old.xlsx",sheet_name='EVENT_4G_RSSI')
df1=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_RSSI')
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//4G_Raw_Total_New.xlsx",float_format="%.2f",index=False)



book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['EVENT_4G_RSSI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_RSSI')
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//4G_Raw_Total_New.xlsx") #Add file name
sheet = wb.worksheets[1] #Add Sheet name
#sheet = wb["EVENT_4G_RSSI"] 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["EVENT_4G_RSSI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,4,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,4,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='EVENT_4G_RSSI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['EVENT_4G_RSSI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(4)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  



########################################################Copy paste 4G Cell data###################################################################
##################################################################################################################################################
        
#################################################################Code to copy paste in Final Output sheet-EVENT_4G_SRVCC_KPI########################################

#df=pd.read_excel(INPATH2+"//4G_Raw_Cell_Old.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')
df1=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 2g][EVENT_4G_SRVCC_KPI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['Cell_4G_SRVCC_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//4G_Raw_Cell_New.xlsx") #Add file name
sheet = wb.worksheets[2] #Add Sheet name
#sheet = wb["EVENT_4G_SRVCC_KPI"] 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["Cell_4G_SRVCC_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,4,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,4,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='Cell_4G_SRVCC_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['Cell_4G_SRVCC_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(4)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
        
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")    


#################################################Copy range and paste in 4g][RSSI_4G_KPI############################################################
#################################################################Code to copy paste in Final Output sheet-Total_4G_RSSI########################################

#df=pd.read_excel(INPATH2+"//4G_Raw_Total_Old.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')
df1=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//4G_Raw_Total_New.xlsx",float_format="%.2f",index=False)



book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['Event_4G_SRVCC_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//4G_Raw_Total_New.xlsx") #Add file name
sheet = wb.worksheets[2] #Add Sheet name
#sheet = wb["EVENT_4G_SRVCC_KPI"] 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["Event_4G_SRVCC_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,3,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,3,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='Event_4G_SRVCC_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['Event_4G_SRVCC_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(3)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  


##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################
####################################################################4G Pasting in the Output sheet is done################################################
##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################


#######################################################################################################################################################
#######################################################################################################################################################
############################################################Copy paste the Raw data to Output file for all Technos and save it#########################
#################################################################Code to copy paste in Final Output sheet-CELL_5G_KPI########################################
#######################################################################################################################################################
#######################################################################################################################################################


#################################################################Code to copy paste in Final Output sheet-CELL_5G_KPI########################################

#df=pd.read_excel(INPATH2+"//3G_Raw_Cell_Old.xlsx")
df1=pd.read_excel(INPATH2+"//5G_Raw_Cell_New.xlsx")
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//3G_Raw_Cell_New_consider.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 2g][CELL_2G_KPI############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['CELL_5G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//5G_Raw_Cell_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//5G_Raw_Cell_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["CELL_5G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,12,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,12,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='CELL_5G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['CELL_5G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(12)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
        
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")    


#################################################################Code to copy paste in Final Output sheet-EVENT_5G_KPI########################################

#df=pd.read_excel(INPATH2+"//3G_Raw_Total_Old.xlsx")
df1=pd.read_excel(INPATH2+"//5G_Raw_Total_New.xlsx")
#db_database=df[['time','Remarks']]
#df1=pd.merge(df1,db_database,on='time',how='outer')
#df2=df1[(df1['Remarks']!='done')]
#df2.to_excel(INPATH2+"//3G_Raw_Total_New_consider.xlsx",float_format="%.2f",index=False)

#################################################Copy range and paste in 3g[CELL_3G_KPI]############################################################
book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
sheet_2G_KPI = book['EVENT_5G_KPI']
lastrow_2G_KPI=len(sheet_2G_KPI['A']) #target last row
print('target last row:lastrow_2G_KPI')
print(lastrow_2G_KPI)

df2=pd.read_excel(INPATH2+"//5G_Raw_Total_New.xlsx")
nrows = len(df2)  #source last row
print('source last row:nrows')
print(nrows)
nrows=nrows+1
lastrow_2G_KPI13=lastrow_2G_KPI+1 #target lastrow+1
print('target lastrow+1:lastrow_2G_KPI13')
print(lastrow_2G_KPI13)
lastrow_2G_KPI11=int(lastrow_2G_KPI)+int(nrows) #to add number of last row in target to copy rows
print(lastrow_2G_KPI11)
print('to add number of last row in target to copy rows:lastrow_2G_KPI11')
 
#File to be copied
wb = openpyxl.load_workbook(INPATH2+"//5G_Raw_Total_New.xlsx") #Add file name
sheet = wb.worksheets[0] #Add Sheet name
 
#File to be pasted into
template = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") #Add file name
temp_sheet = template["EVENT_5G_KPI"] #Add Sheet name
 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow,1):
        countCol = 0
        for j in range(startCol,endCol,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("Processing...")
selectedRange = copyRange(1,2,11,nrows,sheet) #Change the 4 number values
pastingRange = pasteRange(1,lastrow_2G_KPI13,11,lastrow_2G_KPI11,temp_sheet,selectedRange) #Change the 4 number values
#You can save the template as another file to create a new file here too.s'''

template.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx") 
##########################################################apply borders and change font######################################################
df22 = pd.read_excel(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx",sheet_name='EVENT_5G_KPI')

book = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
ws = book['EVENT_5G_KPI']
srow=int(lastrow_2G_KPI13)
mrows = ws.max_row  #max rows
mrows =mrows + 1
scol =int(1)
mcol=int(11)
print(mrows)
for row in range(srow,mrows):
    for column in range(scol,mcol):
        ws.cell(row=row,column=column).font = Font(size=8)
        top_left_cell = ws.cell(row=row,column=column)
        top_left_cell.font  = Font(b=False, color="000000")
        top_left_cell.font  = Font(name="Calibri", sz=8)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(border_style="thin", color="000000")
        top_left_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        top_left_cell.alignment = Alignment(wrap_text=False,vertical='center')
book.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")  




##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################
####################################################################5G Pasting in the Output sheet is done################################################
##########################################################################################################################################################
##########################################################################################################################################################
##########################################################################################################################################################


##############################################################################Add Remarks for New Edited ITK Excel Files###############################

#######################################################2G Remarks################################################
df2g_total=pd.read_excel(INPATH2+"//2G_Raw_Total_New.xlsx")
def comment_datatwo(df2g_total):
    if (df2g_total['time']!=''):    
        return "done"

    else:
        return "not"

df2g_total['Remarks'] = df2g_total.apply(comment_datatwo, axis = 1)
df2g_total.to_excel(INPATH2+"//2G_Raw_Total_New.xlsx",float_format="%.2f",index=False)

#######################################################2G Remarks################################################
df2g_total=pd.read_excel(INPATH2+"//2G_Raw_Cell_New.xlsx")
def comment_datatwo(df2g_total):
    if (df2g_total['time']!=''):    
        return "done"

    else:
        return "not"

df2g_total['Remarks'] = df2g_total.apply(comment_datatwo, axis = 1)
df2g_total.to_excel(INPATH2+"//2G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)

#######################################################3G Remarks################################################
df3g_total=pd.read_excel(INPATH2+"//3G_Raw_Total_New.xlsx")
def comment_datatwo(df3g_total):
    #df2g_total['Remarks']=df2g_total['GSM_TRAFIK_VOLUME_ERLANG']
    #if (df2g_total['GSM_TRAFIK_VOLUME_ERLANG']>0):
    if (df3g_total['time']!=''):    
        return "done"

    else:
        return "not"

df3g_total['Remarks'] = df3g_total.apply(comment_datatwo, axis = 1)
df3g_total.to_excel(INPATH2+"//3G_Raw_Total_New.xlsx",float_format="%.2f",index=False)

#######################################################3G Remarks################################################
df3g_total=pd.read_excel(INPATH2+"//3G_Raw_Cell_New.xlsx")
def comment_datatwo(df3g_total):
    if (df3g_total['time']!=''):    
        return "done"

    else:
        return "not"

df3g_total['Remarks'] = df3g_total.apply(comment_datatwo, axis = 1)
df3g_total.to_excel(INPATH2+"//3G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)


#######################################################4G Remarks################################################
#######################################################4G Remarks################################################
df4g_total=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_KPI')
df5g_total=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_RSSI')
df6g_total=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')

def comment_datatwo(df4g_total):
    if (df4g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datathree(df5g_total):
    if (df5g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datafour(df6g_total):
    if (df6g_total['time']!=''):    
        return "done"

    else:
        return "not"    

df4g_total['Remarks'] = df4g_total.apply(comment_datatwo, axis = 1)
df5g_total['Remarks'] = df5g_total.apply(comment_datathree, axis = 1)
df6g_total['Remarks'] = df6g_total.apply(comment_datafour, axis = 1)


with pd.ExcelWriter(INPATH2+"//4G_Raw_Cell_New.xlsx") as writer:
      
    df4g_total.to_excel(writer,sheet_name='EVENT_4G_KPI',float_format="%.2f",index=False)
    df5g_total.to_excel(writer,sheet_name='EVENT_4G_RSSI',float_format="%.2f",index=False)
    df6g_total.to_excel(writer,sheet_name='EVENT_4G_SRVCC_KPI',float_format="%.2f",index=False)
writer.save()

#######################################################4G Remarks Total################################################
df4g_total=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_KPI')
df5g_total=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_RSSI')
df6g_total=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')

def comment_datatwo(df4g_total):
    if (df4g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datathree(df5g_total):
    if (df5g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datafour(df6g_total):
    if (df6g_total['time']!=''):    
        return "done"

    else:
        return "not"    

df4g_total['Remarks'] = df4g_total.apply(comment_datatwo, axis = 1)
df5g_total['Remarks'] = df5g_total.apply(comment_datathree, axis = 1)
df6g_total['Remarks'] = df6g_total.apply(comment_datafour, axis = 1)


with pd.ExcelWriter(INPATH2+"//4G_Raw_Total_New.xlsx") as writer:
      
    df4g_total.to_excel(writer,sheet_name='EVENT_4G_KPI',float_format="%.2f",index=False)
    df5g_total.to_excel(writer,sheet_name='EVENT_4G_RSSI',float_format="%.2f",index=False)
    df6g_total.to_excel(writer,sheet_name='EVENT_4G_SRVCC_KPI',float_format="%.2f",index=False)
writer.save()

###############################################################Give as done and remane as old###############################################
##############################################################################Add Remarks for New Edited ITK Excel Files###############################

#######################################################2G Remarks################################################
df2g_total=pd.read_excel(INPATH2+"//2G_Raw_Total_New.xlsx")
def comment_datatwo(df2g_total):
    if (df2g_total['time']!=''):    
        return "done"

    else:
        return "not"

df2g_total['Remarks'] = df2g_total.apply(comment_datatwo, axis = 1)
df2g_total.to_excel(INPATH2+"//2G_Raw_Total_New.xlsx",float_format="%.2f",index=False)

#######################################################2G Remarks################################################
df2g_total=pd.read_excel(INPATH2+"//2G_Raw_Cell_New.xlsx")
def comment_datatwo(df2g_total):
    if (df2g_total['time']!=''):    
        return "done"

    else:
        return "not"

df2g_total['Remarks'] = df2g_total.apply(comment_datatwo, axis = 1)
df2g_total.to_excel(INPATH2+"//2G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)

#######################################################3G Remarks################################################
df3g_total=pd.read_excel(INPATH2+"//3G_Raw_Total_New.xlsx")
def comment_datatwo(df3g_total):
    #df2g_total['Remarks']=df2g_total['GSM_TRAFIK_VOLUME_ERLANG']
    #if (df2g_total['GSM_TRAFIK_VOLUME_ERLANG']>0):
    if (df3g_total['time']!=''):    
        return "done"

    else:
        return "not"

df3g_total['Remarks'] = df3g_total.apply(comment_datatwo, axis = 1)
df3g_total.to_excel(INPATH2+"//3G_Raw_Total_New.xlsx",float_format="%.2f",index=False)

#######################################################3G Remarks################################################
df3g_total=pd.read_excel(INPATH2+"//3G_Raw_Cell_New.xlsx")
def comment_datatwo(df3g_total):
    if (df3g_total['time']!=''):    
        return "done"

    else:
        return "not"

df3g_total['Remarks'] = df3g_total.apply(comment_datatwo, axis = 1)
df3g_total.to_excel(INPATH2+"//3G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)


#######################################################4G Remarks################################################
#######################################################4G Remarks################################################
df4g_total=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_KPI')
df5g_total=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_RSSI')
df6g_total=pd.read_excel(INPATH2+"//4G_Raw_Cell_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')

def comment_datatwo(df4g_total):
    if (df4g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datathree(df5g_total):
    if (df5g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datafour(df6g_total):
    if (df6g_total['time']!=''):    
        return "done"

    else:
        return "not"    

df4g_total['Remarks'] = df4g_total.apply(comment_datatwo, axis = 1)
df5g_total['Remarks'] = df5g_total.apply(comment_datathree, axis = 1)
df6g_total['Remarks'] = df6g_total.apply(comment_datafour, axis = 1)


with pd.ExcelWriter(INPATH2+"//4G_Raw_Cell_New.xlsx") as writer:
      
    df4g_total.to_excel(writer,sheet_name='EVENT_4G_KPI',float_format="%.2f",index=False)
    df5g_total.to_excel(writer,sheet_name='EVENT_4G_RSSI',float_format="%.2f",index=False)
    df6g_total.to_excel(writer,sheet_name='EVENT_4G_SRVCC_KPI',float_format="%.2f",index=False)
writer.save()

#######################################################4G Remarks Total################################################
df4g_total=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_KPI')
df5g_total=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_RSSI')
df6g_total=pd.read_excel(INPATH2+"//4G_Raw_Total_New.xlsx",sheet_name='EVENT_4G_SRVCC_KPI')

def comment_datatwo(df4g_total):
    if (df4g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datathree(df5g_total):
    if (df5g_total['time']!=''):    
        return "done"

    else:
        return "not"

def comment_datafour(df6g_total):
    if (df6g_total['time']!=''):    
        return "done"

    else:
        return "not"    

df4g_total['Remarks'] = df4g_total.apply(comment_datatwo, axis = 1)
df5g_total['Remarks'] = df5g_total.apply(comment_datathree, axis = 1)
df6g_total['Remarks'] = df6g_total.apply(comment_datafour, axis = 1)


with pd.ExcelWriter(INPATH2+"//4G_Raw_Total_New.xlsx") as writer:
      
    df4g_total.to_excel(writer,sheet_name='EVENT_4G_KPI',float_format="%.2f",index=False)
    df5g_total.to_excel(writer,sheet_name='EVENT_4G_RSSI',float_format="%.2f",index=False)
    df6g_total.to_excel(writer,sheet_name='EVENT_4G_SRVCC_KPI',float_format="%.2f",index=False)
writer.save()




#######################################################5G Remarks################################################
'''df5g_total=pd.read_excel(INPATH2+"//5G_Raw_Total_New.xlsx")
def comment_datafive(df5g_total):
    if (df5g_total['time']!=''):    
        return "done"

    else:
        return "not"

df5g_total['Remarks'] = df5g_total.apply(comment_datafive, axis = 1)
df5g_total.to_excel(INPATH2+"//5G_Raw_Total_New.xlsx",float_format="%.2f",index=False)
#######################################################5G Remarks################################################
df5g_total=pd.read_excel(INPATH2+"//5G_Raw_Cell_New.xlsx")
def comment_datafive(df5g_total):
    if (df5g_total['time']!=''):    
        return "done"

    else:
        return "not"

df5g_total['Remarks'] = df5g_total.apply(comment_datafive, axis = 1)
df5g_total.to_excel(INPATH2+"//5G_Raw_Cell_New.xlsx",float_format="%.2f",index=False)

##############################################Rename New file as old file################################
file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Cell_New.xlsx")
if file.is_file():
    print ("4G_Raw_Cell_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Cell_Old.xlsx')

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Total_New.xlsx")
if file.is_file():
    print ("4G_Raw_Total_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//4G_Raw_Total_Old.xlsx')

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Cell_New.xlsx")
if file.is_file():
    print ("3G_Raw_Cell_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Cell_Old.xlsx')

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Total_New.xlsx")
if file.is_file():
    print ("3G_Raw_Total_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//3G_Raw_Total_Old.xlsx')

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Cell_New.xlsx")
if file.is_file():
    print ("2G_Raw_Cell_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Cell_Old.xlsx')

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Total_New.xlsx")
if file.is_file():
    print ("2G_Raw_Total_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//2G_Raw_Total_Old.xlsx')

    
file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Cell_New.xlsx")
if file.is_file():
    print ("5G_Raw_Cell_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Cell_Old.xlsx')

file = Path("C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Total_New.xlsx")
if file.is_file():
    print ("5G_Raw_Total_New.xlsx File exist")
    os.rename(file , 'C://EVENTS_AUTO//ITK_Raw_data//ITK_Edited_Data//5G_Raw_Total_Old.xlsx')''''


#################################################################Create Charts##############################################################################
#path="C:\\Users\\ENABGOP\\Documents\\Python Scripts\\Automation_scripts\\chart_prep"
path='C://EVENTS_AUTO//OUTPUT'

wb = openpyxl.load_workbook(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
# ws = wb.active
##2G chart
print(wb.sheetnames)
wg=wb['EVENT_2G_KPI']
wl=wb['EVENT_4G_KPI']
wrss=wb['EVENT_4G_RSSI']
wr=wb['Event_4G_SRVCC_KPI']
wu=wb['EVENT_3G_KPI']
we=wb['EVENT_KPI_GRAPH']
wnr=wb['EVENT_5G_KPI']

from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
font_test = Font(typeface='Calibri')
cp = CharacterProperties(latin=font_test, sz=1500)
from openpyxl.drawing.text import (
    ParagraphProperties,
    CharacterProperties,
)
# def set_chart_title_size(chart, size=1400):
#     paraprops = ParagraphProperties()
#     paraprops.defRPr = CharacterProperties(sz=size)

#     for para in chart.title.tx.rich.paragraphs:
#         para.pPr=paraprops 

# set_chart_title_size(chart, size=1400)
# chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
from openpyxl.drawing.text import CharacterProperties

def GSM(ws,we):
    min_column = ws.min_column
    max_column = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    dates = chart.Reference(ws,
                           min_col=1,
                           max_col=1,
                           min_row=2,
                           max_row=max_row)
    traffic=chart.Reference(ws,
                           min_col=2,
                           max_col=2,
                           min_row=min_row,
                           max_row=max_row)
    Call_Estab=chart.Reference(ws,
                           min_col=4,
                           max_col=4,
                           min_row=min_row,
                           max_row=max_row)

    cdr=chart.Reference(ws,
                           min_col=3,
                           max_col=3,
                           min_row=min_row,
                           max_row=max_row)

    c1 = LineChart()
#     c1.title = "2G_BSS_Call_Estab_SR (%)_with_TRAFFIC(ERL)"
    

    c1.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
   
    #c1.x_axis.title = "Date"
    c1.x_axis.crosses = "min"
    c1.x_axis.majorTickMark = "out"
    c1.x_axis.number_format = "yyyy-dd-mm:hh"

    c1.add_data(traffic, titles_from_data=True)
    c1.set_categories(dates)
    #c1.y_axis.title = 'GSM_TRAFIK_VOLUME_ERLANG'
    c1.y_axis.crossAx = 500
    c1.y_axis.majorGridlines = None
    c1.height=13
    c1.width = 54
    c1.legend.position = 't'
    c1.x_axis.scaling.min = 0 
    c1.y_axis.scaling.min = 0
    
    c2 = LineChart()
    c2.title = "2G_BSS_Call_Estab_SR (%)_with_TRAFFIC(ERL)"
    c2.x_axis.axId = 500 # same as c1

    c2.add_data(Call_Estab, titles_from_data=True)
    c2.set_categories(dates)

    c2.y_axis.axId = 20
    #c2.y_axis.title = "BSS_Call_Estab_SR"
    c2.y_axis.crossAx = 500
    c2.height=13
    c2.width = 54
    c2.legend.position = 't' 
    c2.x_axis.scaling.min = 0
    c2.y_axis.scaling.min = 0
    c1.y_axis.crosses = "max"
    c2 += c1
    
    c3 = LineChart()
    c3.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500

    #c3.x_axis.title = "Date"
    c3.x_axis.crosses = "min"
    c3.x_axis.majorTickMark = "out"
    c3.x_axis.number_format = "yyyy-dd-mm:hh"

    c3.add_data(cdr, titles_from_data=True)
    c3.set_categories(dates)
    #c3.y_axis.title = '2G_CDR'
    c3.y_axis.crossAx = 500
    #c3.y_axis.majorGridlines = None
    c3.x_axis.majorGridlines = None
    c3.x_axis.scaling.min = 0 
    c3.y_axis.scaling.min = 0
    
    c3.height=13
    c3.width = 54
#     c3.plot_area.graphicalProperties = props
    c3.legend.position = 't'

    we.add_chart(c2, "A333")
    we.add_chart(c3, "A358")
    wb.save("Newchart.xlsx")
GSM(wg,we)

def LTE_1(ws,we):
    min_column = ws.min_column
    max_column = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    dates = chart.Reference(ws,
                           min_col=1,
                           max_col=1,
                           min_row=2,
                           max_row=max_row)
    DL_data=chart.Reference(ws,
                           min_col=4,
                           max_col=5,
                           min_row=min_row,
                           max_row=max_row)
#     UL_data=chart.Reference(ws,
#                            min_col=5,
#                            max_col=5,
#                            min_row=min_row,
#                            max_row=max_row)
    LTE_Estab=chart.Reference(ws,
                           min_col=2,
                           max_col=2,
                           min_row=min_row,
                           max_row=max_row)

    psdrop=chart.Reference(ws,
                           min_col=3,
                           max_col=3,
                           min_row=min_row,
                           max_row=max_row)
    DL_Thp=chart.Reference(ws,
                           min_col=6,
                           max_col=6,
                           min_row=min_row,
                           max_row=max_row)
    UL_thp=chart.Reference(ws,
                           min_col=7,
                           max_col=7,
                           min_row=min_row,
                           max_row=max_row)
    DL_user=chart.Reference(ws,
                           min_col=8,
                           max_col=8,
                           min_row=min_row,
                           max_row=max_row)
    UL_user=chart.Reference(ws,
                           min_col=11,
                           max_col=11,
                           min_row=min_row,
                           max_row=max_row)
    Volte_traffic=chart.Reference(ws,
                           min_col=13,
                           max_col=13,
                           min_row=min_row,
                           max_row=max_row)
    Volte_CSSR=chart.Reference(ws,
                           min_col=14,
                           max_col=14,
                           min_row=min_row,
                           max_row=max_row)
    Volte_CDR=chart.Reference(ws,
                           min_col=15,
                           max_col=15,
                           min_row=min_row,
                           max_row=max_row)

    c1 = LineChart()
#     c1.title = "4G_ESTABLISHMENT_SUCCESS_RATE(%)"
    c1.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c1.x_axis.title = "Date"
    c1.x_axis.crosses = "min"
    c1.x_axis.majorTickMark = "out"
    c1.x_axis.number_format = "yyyy-dd-mm:hh"

    c1.add_data(DL_data, titles_from_data=True)
    c1.set_categories(dates)
    #c1.y_axis.title = '4G_DATA_VOLUME_GB'
    c1.y_axis.crossAx = 500
    c1.y_axis.majorGridlines = None

    c1.height=13
    c1.width = 54

    c2 = LineChart()
    c2.x_axis.axId = 500 # same as c1

    c2.add_data(LTE_Estab, titles_from_data=True)
    c2.set_categories(dates)

    c2.y_axis.axId = 20
    c2.title = "4G_ESTABLISHMENT_SUCCESS_RATE(%)"
    #c2.y_axis.title = "4G_ESTABLISHMENT_SUCCESS_RATER"
    c2.y_axis.crossAx = 500
    c2.height=13
    c2.width = 54
    c1.legend.position = 't'
    c2.legend.position = 't'
    c1.y_axis.crosses = "max"
    c2 += c1
    
    
    c3 = LineChart()
    c3.title = "4G_CALL_DROP_RATE"
    c3.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c3.x_axis.title = "Date"
    c3.x_axis.crosses = "min"
    c3.x_axis.majorTickMark = "out"
    c3.x_axis.number_format = "yyyy-dd-mm:hh"

    c3.add_data(psdrop, titles_from_data=True)
    c3.set_categories(dates)
#     c3.y_axis.title = '4G_CALL_DROP_RATE'
    c3.y_axis.crossAx = 500
#     c3.y_axis.majorGridlines = None
    c3.x_axis.majorGridlines = None
    
    c3.height=13
    c3.width = 54
#     c3.plot_area.graphicalProperties = props
    c3.legend.position = 't'
    
        
    c5 = LineChart()
    #c5.title = "4G_USER_THROUGHPUT(Mbps"
    c5.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c5.x_axis.title = "Date"
    c5.x_axis.crosses = "min"
    c5.x_axis.majorTickMark = "out"
    c5.x_axis.number_format = "yyyy-dd-mm:hh"

    c5.add_data(UL_thp, titles_from_data=True)
    c5.set_categories(dates)
    #c5.y_axis.title = 'AVG_DL_USER_THP_Mbps'
    c5.y_axis.crossAx = 500
    c5.y_axis.majorGridlines = None
    c5.height=13
    c5.width = 54
    c5.legend.position = 't'
    
    c6 = LineChart()
    c6.title = "4G_USER_THROUGHPUT(Mbps"
    c6.x_axis.axId = 500 # same as c5

    c6.add_data(DL_Thp, titles_from_data=True)
    c6.set_categories(dates)

    c6.y_axis.axId = 20
    #c6.y_axis.title = "AVG_UL_USER_THP_Mbps"
    c6.y_axis.crossAx = 500
    c6.height=13
    c6.width = 54
    c6.legend.position = 't'
    c5.y_axis.crosses = "max"
    c6 += c5

    c7 = LineChart()
    #c7.title = "4G_DATA_USER_NUM"
    c7.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c7.x_axis.title = "Date"
    c7.x_axis.crosses = "min"
    c7.x_axis.majorTickMark = "out"
    c7.x_axis.number_format = "yyyy-dd-mm:hh"

    c7.add_data(UL_user, titles_from_data=True)
    c7.set_categories(dates)
    #c7.y_axis.title = 'AVG_DATA_USER_NUM_DL'
    c7.y_axis.crossAx = 500
    c7.y_axis.majorGridlines = None
    c7.height=13
    c7.width = 54
    c7.legend.position = 't'
    c8 = LineChart()
    c8.title = "4G_DATA_USER_NUM"
    c8.x_axis.axId = 500 # same as c7

    c8.add_data(DL_user, titles_from_data=True)
    c8.set_categories(dates)

    c8.y_axis.axId = 20
    #c8.y_axis.title = "AVG_DATA_USER_NUM_UL"
    c8.y_axis.crossAx = 500
    c8.height=13
    c8.width = 54
    c8.legend.position = 't'
    c7.y_axis.crosses = "max"
    c8 += c7
    
    c9 = LineChart()
    #c9.title = "VOLTE_CSSR(%)_with_TRAFFIC(ERL)"
    c9.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c9.x_axis.title = "Date"
    c9.x_axis.crosses = "min"
    c9.x_axis.majorTickMark = "out"
    c9.x_axis.number_format = "yyyy-dd-mm:hh"

    c9.add_data(Volte_traffic, titles_from_data=True)
    c9.set_categories(dates)
    #c9.y_axis.title = 'VOLTE_TRAFFIC_ERLANG'
    c9.y_axis.crossAx = 500
    c9.y_axis.majorGridlines = None
    c9.height=13
    c9.width = 54
    c9.legend.position = 't'
    c10 = LineChart()
    c10.title = "VOLTE_CSSR(%)_with_TRAFFIC(ERL)"
    c10.x_axis.axId = 500 # same as c9

    c10.add_data(Volte_CSSR, titles_from_data=True)
    c10.set_categories(dates)

    c10.y_axis.axId = 20
    #c10.y_axis.title = "CSSR_VOLTE"
    c10.y_axis.crossAx = 500
    c10.height=13
    c10.width = 54
    c10.legend.position = 't'
    c9.y_axis.crosses = "max"
    c10 += c9
    
    c11 = LineChart()
    c11.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c11.x_axis.title = "Date"
    c11.x_axis.crosses = "min"
    c11.x_axis.majorTickMark = "out"
    c11.x_axis.number_format = "yyyy-dd-mm:hh"

    c11.add_data(Volte_CDR, titles_from_data=True)
    c11.set_categories(dates)
    #c11.y_axis.title = 'VOLTE_CDR'
    c11.y_axis.crossAx = 500
#     c11.y_axis.majorGridlines = None
    c11.x_axis.majorGridlines = None
    c11.height=13
    c11.width = 54
    c11.legend.position = 't'
    
    we.add_chart(c2, "A4")
    we.add_chart(c3, "A29")
    we.add_chart(c6, "A54")
    we.add_chart(c8, "A79")
    we.add_chart(c10, "A104")
    we.add_chart(c11,"A129")
    
    wb.save("Newchart.xlsx")
    
LTE_1(wl,we)

def LTE_2(ws,we):
    min_column = ws.min_column
    max_column = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    dates = chart.Reference(ws,
                           min_col=1,
                           max_col=1,
                           min_row=2,
                           max_row=max_row)
    PUCCH_RSSI=chart.Reference(ws,
                           min_col=2,
                           max_col=3,
                           min_row=min_row,
                           max_row=max_row)
#     PUSCH_RSSI=chart.Reference(ws,
#                            min_col=3,
#                            max_col=3,
#                            min_row=min_row,
#                            max_row=max_row)
    c12 = LineChart()
    c12.title = "4G_RSSI(dBm)"
    c12.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
   # c12.x_axis.title = "Date"
    c12.x_axis.crosses = "min"
    c12.x_axis.majorTickMark = "out"
    c12.x_axis.number_format = "yyyy-dd-mm:hh"

    c12.add_data(PUCCH_RSSI, titles_from_data=True)
    c12.set_categories(dates)
#     c12.y_axis.title = '4G_RSSI(dBm)'
    c12.y_axis.crossAx = 500
#     c12.y_axis.majorGridlines = None
    c12.height=13
    c12.width = 54
    c12.legend.position = 't'
    c12.x_axis.majorGridlines = None

    
    we.add_chart(c12,"A154")
    
    wb.save("Newchart.xlsx")
    
LTE_2(wrss,we)  

def LTE_3(ws,we):
    min_column = ws.min_column
    max_column = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    dates = chart.Reference(ws,
                           min_col=1,
                           max_col=1,
                           min_row=2,
                           max_row=max_row)
    srvcc=chart.Reference(ws,
                           min_col=2,
                           max_col=3,
                           min_row=min_row,
                           max_row=max_row)
#     PUSCH_RSSI=chart.Reference(ws,
#                            min_col=3,
#                            max_col=3,
#                            min_row=min_row,
#                            max_row=max_row)
    c13 = LineChart()
    c13.title = "VOLTE_SRVCC_HO_SR(%)"
    c13.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c13.x_axis.title = "Date"
    c13.x_axis.crosses = "min"
    c13.x_axis.majorTickMark = "out"
    c13.x_axis.number_format = "yyyy-dd-mm:hh"

    c13.add_data(srvcc, titles_from_data=True)
    c13.set_categories(dates)
#     c13.y_axis.title = '4G_RSSI(dBm)'
    c13.y_axis.crossAx = 500
#     c13.y_axis.majorGridlines = None
    c13.height=13
    c13.width = 54
    c13.legend.position = 't'
    c13.x_axis.majorGridlines = None
    
    we.add_chart(c13,"A178")
    
    wb.save("Newchart.xlsx")
    
LTE_3(wr,we)  

from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.shapes import GraphicalProperties
# props = GraphicalProperties(solidFill="#F0F8FF") 
def UMTS(ws,we):
    min_column = ws.min_column
    max_column = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    dates = chart.Reference(ws,
                           min_col=1,
                           max_col=1,
                           min_row=2,
                           max_row=max_row)
    CSSR=chart.Reference(ws,
                           min_col=2,
                           max_col=2,
                           min_row=min_row,
                           max_row=max_row)
    Traffic=chart.Reference(ws,
                           min_col=3,
                           max_col=3,
                           min_row=min_row,
                           max_row=max_row)
    PS_cssr=chart.Reference(ws,
                           min_col=4,
                           max_col=4,
                           min_row=min_row,
                           max_row=max_row)

    UL_data=chart.Reference(ws,
                           min_col=5,
                           max_col=6,
                           min_row=min_row,
                           max_row=max_row)
#     DL_Thp=chart.Reference(ws,
#                            min_col=6,
#                            max_col=6,
#                            min_row=min_row,
#                            max_row=max_row)
    cdr=chart.Reference(ws,
                           min_col=7,
                           max_col=7,
                           min_row=min_row,
                           max_row=max_row)
    hsdpa_thp=chart.Reference(ws,
                           min_col=9,
                           max_col=9,
                           min_row=min_row,
                           max_row=max_row)
    hsupa_thp=chart.Reference(ws,
                           min_col=10,
                           max_col=10,
                           min_row=min_row,
                           max_row=max_row)
    rtwp=chart.Reference(ws,
                           min_col=8,
                           max_col=8,
                           min_row=min_row,
                           max_row=max_row)
#     props = GraphicalProperties(solidFill="8BADD9") 
    c1 = LineChart()
    #c1.title = "3G_CS_CSSR(%)_with_TRAFFIC(ERL)"
    c1.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c1.x_axis.title = "Date"
    c1.x_axis.crosses = "min"
    c1.x_axis.majorTickMark = "out"
    c1.x_axis.number_format = "yyyy-dd-mm:hh"

    c1.add_data(Traffic, titles_from_data=True)
    c1.set_categories(dates)
#     c1.y_axis.title = '3G_CS_CSSR'
    c1.y_axis.crossAx = 500
    c1.y_axis.majorGridlines = None
    c1.height=13
    c1.width = 54
     
#     c1.plot_area.graphicalProperties = props
    
    c2 = LineChart()
    c2.title = "3G_CS_CSSR(%)_with_TRAFFIC(ERL)"
    c2.x_axis.axId = 500 # same as c1

    c2.add_data(CSSR, titles_from_data=True)
    c2.set_categories(dates)

    c2.y_axis.axId = 20
#     c2.y_axis.title = "3G_TRAFIK_VOLUME_ERLANG"
    c2.y_axis.crossAx = 500
    c2.height=13
    c2.width = 54
    c1.legend.position = 't'
    c2.legend.position = 't'
   
    c1.y_axis.crosses = "max"
    c2 += c1
    

    c5 = LineChart()
    #c5.title = "3G_PS_CSSR(%)_with_DATA(MB)"
    c5.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c5.x_axis.title = "Date"
    c5.x_axis.crosses = "min"
    c5.x_axis.majorTickMark = "out"
    c5.x_axis.number_format = "yyyy-dd-mm:hh"

    c5.add_data(UL_data, titles_from_data=True)
    c5.set_categories(dates)
#     c5.y_axis.title = 'AVG_DL_USER_THP_Mbps'
    c5.y_axis.crossAx = 500
    c5.y_axis.majorGridlines = None

    c6 = LineChart()
    c6.title = "3G_PS_CSSR(%)_with_DATA(MB)"
    c6.x_axis.axId = 500 # same as c5

    c6.add_data(PS_cssr, titles_from_data=True)
    c6.set_categories(dates)

    c6.y_axis.axId = 20
#     c6.y_axis.title = "AVG_UL_USER_THP_Mbps"
    c6.y_axis.crossAx = 500
    c5.height=13
    c5.width = 54
    c6.height=13
    c6.width = 54
#     c5.plot_area.graphicalProperties = props
#     c6.plot_area.graphicalProperties = props
    c5.legend.position = 't'
    c6.legend.position = 't'
    c5.y_axis.crosses = "max"
    c6 += c5

    
    c3 = LineChart()
    c3.title = "3G_CS_CDR(%)"
    c3.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c3.x_axis.title = "Date"
    c3.x_axis.crosses = "min"
    c3.x_axis.majorTickMark = "out"
    c3.x_axis.number_format = "yyyy-dd-mm:hh"

    c3.add_data(cdr, titles_from_data=True)
    c3.set_categories(dates)
#     c3.y_axis.title = '4G_CALL_DROP_RATE'
    c3.y_axis.crossAx = 500
#     c3.y_axis.majorGridlines = None
    c3.x_axis.majorGridlines = None
    
    c3.height=13
    c3.width = 54
#     c3.plot_area.graphicalProperties = props
    c3.legend.position = 't'
    
    c9 = LineChart()
    #c9.title = "3G_USER_THROUGHPUT(Mbps)"
    c9.x_axis = chart.axis.DateAxis(crossAx=500) # axId defaults to 500
    #c9.x_axis.title = "Date"
    c9.x_axis.crosses = "min"
    c9.x_axis.majorTickMark = "out"
    c9.x_axis.number_format = "yyyy-dd-mm:hh"

    c9.add_data(hsupa_thp, titles_from_data=True)
    c9.set_categories(dates)
#     c9.y_axis.title = '3G_USER_HSDPA_THROUGHPUT_Kbps'
    c9.y_axis.crossAx = 500
    c9.y_axis.majorGridlines = None

    c10 = LineChart()
    c10.title = "3G_USER_THROUGHPUT(Mbps)"
    c10.x_axis.axId = 500 # same as c9

    c10.add_data(hsdpa_thp, titles_from_data=True)
    c10.set_categories(dates)

    c10.y_axis.axId = 20
#     c10.y_axis.title = "3G_USER_HSUPA_THROUGHPUT_Kbps"
    c10.y_axis.crossAx = 500

    c9.y_axis.crosses = "max"
    c10 += c9
    
    c9.height=13
    c9.width = 54
    
    c10.height=13
    c10.width = 54
    
#     c9.plot_area.graphicalProperties = props

    c9.legend.position = 't'
    c10.legend.position = 't'

    c13 = LineChart()
    c13.title = "3G_RTWP(dBm)"
    c13.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c13.x_axis.title = "Date"
    c13.x_axis.crosses = "min"
    c13.x_axis.majorTickMark = "out"
    c13.x_axis.number_format = "yyyy-dd-mm:hh"

    c13.add_data(rtwp, titles_from_data=True)
    c13.set_categories(dates)
#     c13.y_axis.title = '4G_RSSI(dBm)'
    c13.y_axis.crossAx = 500
#     c13.y_axis.majorGridlines = None
    c13.x_axis.majorGridlines = None
    c13.height=13
    c13.width = 54
#     c13.plot_area.graphicalProperties = props
    c13.legend.position = 't'
    we.add_chart(c2,"A205")
    we.add_chart(c6,"A230")
    we.add_chart(c3,"A255")
    we.add_chart(c10,"A280")
    we.add_chart(c13,"A305")
    wb.save("Newchart.xlsx")
UMTS(wu,we)


def NR(ws,we):
    min_column = ws.min_column
    max_column = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    dates = chart.Reference(ws,
                           min_col=1,
                           max_col=1,
                           min_row=2,
                           max_row=max_row)
    avail=chart.Reference(ws,
                           min_col=2,
                           max_col=2,
                           min_row=min_row,
                           max_row=max_row)
    estb=chart.Reference(ws,
                           min_col=3,
                           max_col=3,
                           min_row=min_row,
                           max_row=max_row)
    cdr=chart.Reference(ws,
                           min_col=4,
                           max_col=4,
                           min_row=min_row,
                           max_row=max_row)

    rssi=chart.Reference(ws,
                           min_col=6,
                           max_col=6,
                           min_row=min_row,
                           max_row=max_row)
    dl_Thp=chart.Reference(ws,
                            min_col=7,
                            max_col=7,
                            min_row=min_row,
                            max_row=max_row)
    ul_Thp=chart.Reference(ws,
                            min_col=8,
                            max_col=8,
                            min_row=min_row,
                            max_row=max_row)

    dl_traffic=chart.Reference(ws,
                           min_col=9,
                           max_col=9,
                           min_row=min_row,
                           max_row=max_row)
    ul_traffic=chart.Reference(ws,
                           min_col=10,
                           max_col=10,
                           min_row=min_row,
                           max_row=max_row)
   
    c3 = LineChart()
    c3.title = "5G_Availability"
    c3.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c3.x_axis.title = "Date"
    c3.x_axis.crosses = "min"
    c3.x_axis.majorTickMark = "out"
    c3.x_axis.number_format = "yyyy-dd-mm:hh"

    c3.add_data(avail, titles_from_data=True)
    c3.set_categories(dates)

    c3.y_axis.crossAx = 500
    c3.x_axis.majorGridlines = None
    c3.height=13
    c3.width = 54
    c3.legend.position = 't'


    c4 = LineChart()
    c4.title = "5G_EN_DC_Est_SR"
    c4.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c4.x_axis.title = "Date"
    c4.x_axis.crosses = "min"
    c4.x_axis.majorTickMark = "out"
    c4.x_axis.number_format = "yyyy-dd-mm:hh"

    c4.add_data(estb, titles_from_data=True)
    c4.set_categories(dates)

    c4.y_axis.crossAx = 500
    c4.x_axis.majorGridlines = None
    c4.height=13
    c4.width = 54
    c4.legend.position = 't'


    c5 = LineChart()
    c5.title = "5G_EN_DC_CDR"
    c5.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c5.x_axis.title = "Date"
    c5.x_axis.crosses = "min"
    c5.x_axis.majorTickMark = "out"
    c5.x_axis.number_format = "yyyy-dd-mm:hh"

    c5.add_data(cdr, titles_from_data=True)
    c5.set_categories(dates)

    c5.y_axis.crossAx = 500
    c5.x_axis.majorGridlines = None
    c5.height=13
    c5.width =54
    c5.legend.position = 't'


    c6 = LineChart()
    c6.title = "5G_Average_Uplink_Interference_dBm"
    c6.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c6.x_axis.title = "Date"
    c6.x_axis.crosses = "min"
    c6.x_axis.majorTickMark = "out"
    c6.x_axis.number_format = "yyyy-dd-mm:hh"

    c6.add_data(rssi, titles_from_data=True)
    c6.set_categories(dates)

    c6.y_axis.crossAx = 500
    c6.x_axis.majorGridlines = None
    c6.height=13
    c6.width = 54
    c6.legend.position = 't'


    c7 = LineChart()
    #c7.title = "5G Average User DL & UL Thp Mbps"
    c7.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c7.x_axis.title = "Date"
    c7.x_axis.crosses = "min"
    c7.x_axis.majorTickMark = "out"
    c7.x_axis.number_format = "yyyy-dd-mm:hh"

    c7.add_data(ul_Thp, titles_from_data=True)
    c7.set_categories(dates)
    c7.y_axis.crossAx = 500
    c7.y_axis.majorGridlines = None

    c8 = LineChart()
    c8.title = "5G Average User DL & UL Thp Mbps"
    c8.x_axis.axId = 500 # same as c7
    c8.add_data(dl_Thp, titles_from_data=True)
    c8.set_categories(dates)
    c8.y_axis.axId = 20
    c8.y_axis.crossAx = 500
    c7.height=13
    c7.width = 54
    c8.height=13
    c8.width = 54
    c7.legend.position = 't'
    c8.legend.position = 't'
    c7.y_axis.crosses = "max"
    c8 += c7

    
    c9 = LineChart()
    #c9.title = "5G DL & UL MAC Traffic Volume_GB"
    c9.x_axis = chart.axis.DateAxis(crossAx=100) # axId defaults to 500
    #c9.x_axis.title = "Date"
    c9.x_axis.crosses = "min"
    c9.x_axis.majorTickMark = "out"
    c9.x_axis.number_format = "yyyy-dd-mm:hh"

    c9.add_data(ul_traffic, titles_from_data=True)
    c9.set_categories(dates)
    c9.y_axis.crossAx = 500
    c9.y_axis.majorGridlines = None

    c10 = LineChart()
    c10.title = "5G DL & UL MAC Traffic Volume_GB"
    c10.x_axis.axId = 500 # same as c9
    c10.add_data(dl_traffic, titles_from_data=True)
    c10.set_categories(dates)
    c10.y_axis.axId = 20
    c10.y_axis.crossAx = 500
    c9.height=13
    c9.width = 54
    c10.height=13
    c10.width = 54
    c9.legend.position = 't'
    c10.legend.position = 't'
    c9.y_axis.crosses = "max"
    c10 += c9


    we.add_chart(c3,"A385")
    we.add_chart(c4,"A410")
    we.add_chart(c5,"A435")
    we.add_chart(c6,"A460")
    we.add_chart(c8,"A485")
    we.add_chart(c10,"A510")


    wb.save(OUTPATH+"//EVENT EM 2021_V2_ITK.xlsx")
NR(wnr,we)

